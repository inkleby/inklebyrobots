'''
QuickList library - very simple communication with spreadsheets. 
v1
'''
import csv
import ucsv
import os

try:
    from openpyxl import load_workbook
    import xlrd
    import xlwt
    from xlwt import Workbook
except:
    load_workbook = None
    xlrd = None
    xlwt = None
    Workbook = object

class FlexiBook(Workbook):
    """
    modification to xlwt library - for merging multiple sheets
    """
    def __init__(self,*args,**kwargs):
        super(FlexiBook,self).__init__(*args,**kwargs)

    def add_sheet_from_ql(self,ql):

        sheet_name = ql.name
        if not sheet_name:
            sheet_name = "quicklist"
        ws = self.add_sheet(sheet_name,True)

        row = 0

        for i, h in enumerate(ql.header):
            ws.write(row, i, h)

        for r in ql.data:
            row += 1
            for iter, item in enumerate(r):
                ws.write(row, iter, item)

        return self
    
    def merge_qls(self,qls):
        for q in qls:
            self.add_sheet_from_ql(q)
            
        return self

        

def import_csv(s_file,use_unicode=False):
    """
    imports csvs, returns head/body - switch to use different csv processor
    """
    body = []
    if use_unicode:
        mod = ucsv
    else:
        mod = csv
    
    
    def fix_bom(row):
        nrow = []
        for u in row:
            if type(u) in [str,unicode] and u.startswith(u'\ufeff'):
                u = u[1:]
            nrow.append(u)
        return nrow
    
    with open(s_file, 'rb') as f:
        reader = mod.reader(f)
        for row in reader:
            if unicode:
                body.append(fix_bom(row))
            else:
                body.append(row)

    header = body[0]
    body.pop(0)
    return header, body




def import_xls(s_file, tab="", header_row=0):
    """
    does what you'd expect
    """
    wb = xlrd.open_workbook(filename=s_file)
    if tab == "":
        tab = 0
    if type(tab) == int:
        ws = wb.sheet_by_index(tab)
    else:
        ws = wb.sheet_by_name(tab)  # ws is now an IterableWorksheet

    def generate_rows():
        num_cols = ws.ncols
        for r in range(0, ws.nrows):
            row = [ws.cell(r, c) for c in range(0, num_cols)]
            yield row

    header = []
    data = []
    for count, row in enumerate(generate_rows()):
        r = [x.value for x in row]
        if count == header_row:
            header = r
        if count > header_row:
            data.append(r)

    return header, data


def import_xlsx(s_file, tab="", header_row=0):
    """
    imports xlsx file
    """
    wb = load_workbook(filename=s_file, read_only=True)
    if tab == "":
        tab = wb.sheetnames[0]
    ws = wb[tab]  # ws is now an IterableWorksheet

    header = []
    data = []
    for count, row in enumerate(ws.rows):
        r = [x.value for x in row]
        if count == header_row:
            header = r
        if count > header_row:
            data.append(r)

    return header, data


def export_csv(file, header, body, force_unicode=False):
    """
    exports csv (non-unicode)
    """
    
    if force_unicode:
        module = ucsv
    else:
        module = csv
    with open(file, 'wb') as f:
        w = module.writer(f)
        w.writerows([header])
        w.writerows(body)

class QuickList(object):

    """
    A very simple file interface - loads file into memory so basic reads can be done. 
    """
    
    def __init__(self, name=""):

        self.name = name
        self.header = []
        self.data = []
        self.filename = None

    def union(self,list_of_qls):
        """
        merge identical header files
        """
        
        if not self.header:
            self.header = list_of_qls[0].header
            
        for l in list_of_qls:
            self.data.extend(l.data)
            
        return self
        

    def add(self,row):
        try:
            assert len(row) == len(self.header)
        except Exception, e:
            raise ValueError("New row is larger than header.")
        self.data.append(row)

    def only(self,col,value):
        """
        returns an generator of only rows where col = value
        """
        def make_safe(v):
            if v:
                return v.lower().strip()
            else:
                return ""
            
        safe_col = make_safe(col)
        
        di = self.header_di()
        location = di[safe_col]
        indexes = [x for x,y in enumerate(self.data) if y[location] == value]
        
        return self.__iter__(indexes=indexes)
        
    def exclude(self,col,value):
        """
        returns an generator of only rows where col != value
        """
        def make_safe(v):
            if v:
                return v.lower().strip()
            else:
                return ""
            
        safe_col = make_safe(col)
        
        di = self.header_di()
        location = di[safe_col]
        indexes = [x for x,y in enumerate(self.data) if y[location] != value]
        
        return self.__iter__(indexes=indexes)
        


    def unique(self,col):
        """
        returns an generator of rows with unique values in col
        """
        def make_safe(v):
            if v:
                return v.lower().strip()
            else:
                return ""
            
        safe_col = make_safe(col)
        
        di = self.header_di()
        
        location = di[safe_col]
        
        already_used = []
        
        indexes = []
        
        for x,r in enumerate(self.data):
            o = r[location]
            h = hash(o)
            if h in already_used:
                continue
            else:
                indexes.append(x)
                already_used.append(h)
        
        return self.__iter__(indexes=indexes)
        

    def xls_book(self):
        """
        create xls book from current ql
        """
        wb = FlexiBook()
        wb.add_sheet_from_ql(self)
        return wb

    def open(self, filename, tab="", header_row=0,use_unicode=False):
        """
        populate this ql from a file
        """
        print "Opening : {0}".format(filename)
        ext = os.path.splitext(filename)[1]
        if ext == ".xlsx":
            self.header, self.data = import_xlsx(filename, tab, header_row)
        elif ext == ".csv":
            self.header, self.data = import_csv(filename,unicode)
        elif ext == ".xls":
            self.header, self.data = import_xls(filename, tab, header_row)
        self.filename = filename
        return self

    def save(self, filename=None, force_unicode=False):
        """
        save out as a csv or xls
        """
        print "Saving : {0}".format(filename)
        if filename:
            file_to_use = filename
        else:
            file_to_use = self.filename
            
        if ".csv" in file_to_use:
            export_csv(file_to_use, self.header, self.data, force_unicode=force_unicode)
        if ".xls" in file_to_use:
            self.xls_book().save(file_to_use)
        

    def use_query(self,query):
        """
        use the header to populate with information out of a django query
        """
        self.data = [[y for y in x] for x in query.values_list(*self.header)]
        return self

    def header_di(self):
        """
        returns a dictionary of 'safe' header (striped, lowered) and positions. 
        """
        def make_safe(v):
            if v:
                return v.lower().strip().replace('\ufeff',"")
            else:
                return ""

        return {make_safe(h): x for x, h in enumerate(self.header)}


    def shuffle(self):
        import random
        random.shuffle(self.data)
         
    def construct_item_row_class(self):
        
        header_dict = self.header_di()

        class ItemRow(list):

            def _key_to_index(self, value):
                r_v = value.strip().lower()
                try:
                    lookup = header_dict[r_v]
                except KeyError:
                    lookup = value
                return lookup

            def __getitem__(self, key):
                return super(ItemRow, self).__getitem__(self._key_to_index(key))

            def __setitem__(self, key, value):
                return super(ItemRow, self).__setitem__(self._key_to_index(key), value)
            
        return ItemRow
            

    def get_row(self,row_id):
        
        ItemRow = self.construct_item_row_class()
        
        return ItemRow(self.data[row_id])

    def __iter__(self,indexes=None):
        """
        generator that returns an object that's
         __getitem__ accepts the column name to return a cell instance.
        
        so you can go:
        
        for r in ql:
            p = r["person"]
            
        to get the info out of the person column
         
        """

        ItemRow = self.construct_item_row_class()

        # puts back any changes into the main generator to update the
        # QuickList 
        for x,r in enumerate(self.data):
            if indexes and x not in indexes: # if only yielding certain things, ignore all others
                continue
            ir = ItemRow(r)
            yield ir
            r[:] = ir[:]
            
